"""峰值亮度测试用例生成器 - 独立 Windows GUI 工具

通过界面输入参数生成测试用例 Excel 文件。
迭代顺序：列表从上到下 = 最外层(变化最慢/主分组) → 最内层(变化最快)。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────── 样式常量 ────────────────────

BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

DEFAULT_MODE_COLORS = {
    "标准": "E2EFDA",
    "影院": "D6E4F0",
    "鲜艳": "FFF2CC",
}

PRESET_COLORS = {
    "浅绿": "E2EFDA", "浅蓝": "D6E4F0", "浅黄": "FFF2CC",
    "浅橙": "FCE4D6", "浅紫": "E4DFEC", "浅灰": "F2F2F2",
    "深绿": "C6E0B4", "深蓝": "9DC3E6", "深橙": "F8CBAD",
}

# 因子定义：key, 中文名, 默认选项
FACTOR_DEFS = [
    ("picture_mode", "图像模式", ["标准", "影院", "鲜艳"]),
    ("peak_brightness", "峰值亮度", ["高", "中", "低"]),
    ("contrast_enhance", "对比度增强", ["关", "低", "中", "强"]),
    ("local_dimming", "Local Dimming", ["强"]),
    ("hdr_sdr", "HDR/SDR", ["HDR", "SDR"]),
]

# 迭代顺序：所有可选因子标签 → key 映射
ITER_LABEL_TO_KEY = {
    "图像模式": "picture_mode",
    "峰值亮度": "peak_brightness",
    "对比度增强": "contrast_enhance",
    "Local Dimming": "local_dimming",
    "HDR/SDR": "hdr_sdr",
    "白块亮度": "brightness",
    "窗口大小": "window",
}

# 默认迭代顺序：从上(最外/主分组)到下(最内/变化最快)
DEFAULT_ITER_ORDER = ["图像模式", "峰值亮度", "对比度增强", "Local Dimming", "HDR/SDR", "白块亮度", "窗口大小"]


# ──────────────────── 参数面板 ────────────────────

class FactorFrame(ttk.LabelFrame):
    """一个因子的输入区域：勾选框列表 + 自定义输入。"""

    def __init__(self, parent, label, options, on_change):
        super().__init__(parent, text=label, padding=4)
        self._on_change = on_change
        self._vars: dict[str, tk.BooleanVar] = {}

        for opt in options:
            var = tk.BooleanVar(value=True)
            self._vars[opt] = var
            cb = ttk.Checkbutton(self, text=opt, variable=var, command=self._changed)
            cb.pack(side=tk.LEFT, padx=2)

        ttk.Label(self, text="自定义:").pack(side=tk.LEFT, padx=(8, 2))
        self._custom_entry = ttk.Entry(self, width=20)
        self._custom_entry.pack(side=tk.LEFT, padx=2)
        self._custom_entry.bind("<KeyRelease>", lambda e: self._changed())

    def _changed(self):
        if self._on_change:
            self._on_change()

    def get_values(self) -> list[str]:
        result = [opt for opt, var in self._vars.items() if var.get()]
        custom = self._custom_entry.get().strip()
        if custom:
            for v in custom.replace("，", ",").split(","):
                v = v.strip()
                if v and v not in result:
                    result.append(v)
        return result


# ──────────────────── 主窗口 ────────────────────

class GeneratorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("峰值亮度测试用例生成器")
        root.geometry("960x720")
        root.minsize(800, 600)

        # ── 顶部参数区 ──
        top = ttk.Frame(root, padding=6)
        top.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(top)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 固定参数
        fixed_frame = ttk.LabelFrame(left, text="固定参数", padding=6)
        fixed_frame.pack(fill=tk.X, pady=(0, 4))

        # 因子面板
        self._factor_frames: dict[str, FactorFrame] = {}
        for key, label, options in FACTOR_DEFS:
            frame = FactorFrame(left, label, options, on_change=self._refresh_preview)
            frame.pack(fill=tk.X, pady=2)
            self._factor_frames[key] = frame

        # 白块亮度 — HDR / SDR 各一组
        bright_frame = ttk.LabelFrame(left, text="白块亮度(nit) — HDR / SDR 可分别设置", padding=6)
        bright_frame.pack(fill=tk.X, pady=2)

        self._bright_vars: dict[str, dict] = {}
        for row_idx, (label, defaults) in enumerate([
            ("HDR", ("128", "200", "", "144,167")),
            ("SDR", ("120", "200", "", "144,167")),
        ]):
            start_var = tk.StringVar(value=defaults[0])
            end_var = tk.StringVar(value=defaults[1])
            step_var = tk.StringVar(value=defaults[2])
            extra_var = tk.StringVar(value=defaults[3])

            ttk.Label(bright_frame, text=f"{label}:").grid(row=row_idx, column=0, padx=2, sticky=tk.W)
            ttk.Label(bright_frame, text="起始").grid(row=row_idx, column=1, padx=1)
            ttk.Entry(bright_frame, textvariable=start_var, width=5).grid(row=row_idx, column=2, padx=1)
            ttk.Label(bright_frame, text="结束").grid(row=row_idx, column=3, padx=1)
            ttk.Entry(bright_frame, textvariable=end_var, width=5).grid(row=row_idx, column=4, padx=1)
            ttk.Label(bright_frame, text="间隔").grid(row=row_idx, column=5, padx=1)
            ttk.Entry(bright_frame, textvariable=step_var, width=5).grid(row=row_idx, column=6, padx=1)
            ttk.Label(bright_frame, text="额外").grid(row=row_idx, column=7, padx=1)
            ttk.Entry(bright_frame, textvariable=extra_var, width=10).grid(row=row_idx, column=8, padx=1)

            for w in (start_var, end_var, step_var):
                w.trace_add("write", lambda *_: self._refresh_preview())
            extra_var.trace_add("write", lambda *_: self._refresh_preview())

            self._bright_vars[label] = {
                "start": start_var, "end": end_var,
                "step": step_var, "extra": extra_var,
            }

        # 窗口大小
        win_frame = ttk.LabelFrame(left, text="窗口大小(%)", padding=6)
        win_frame.pack(fill=tk.X, pady=2)

        ttk.Label(win_frame, text="起始:").grid(row=0, column=0, padx=2)
        self._win_start = tk.StringVar(value="0")
        ttk.Entry(win_frame, textvariable=self._win_start, width=5).grid(row=0, column=1, padx=2)

        ttk.Label(win_frame, text="结束:").grid(row=0, column=2, padx=2)
        self._win_end = tk.StringVar(value="100")
        ttk.Entry(win_frame, textvariable=self._win_end, width=5).grid(row=0, column=3, padx=2)

        ttk.Label(win_frame, text="间隔:").grid(row=0, column=4, padx=2)
        self._win_step = tk.StringVar(value="5")
        ttk.Entry(win_frame, textvariable=self._win_step, width=5).grid(row=0, column=5, padx=2)

        ttk.Label(win_frame, text="额外:").grid(row=0, column=6, padx=(8, 2))
        self._win_extra = ttk.Entry(win_frame, width=10)
        self._win_extra.grid(row=0, column=7, padx=2)
        self._win_extra.insert(0, "")

        for w in (self._win_start, self._win_end, self._win_step):
            w.trace_add("write", lambda *_: self._refresh_preview())
        self._win_extra.bind("<KeyRelease>", lambda e: self._refresh_preview())

        # 迭代顺序 — Listbox + 上/下按钮
        iter_frame = ttk.LabelFrame(left, text="迭代顺序（上=最外/主分组，下=最内/变化最快）", padding=6)
        iter_frame.pack(fill=tk.X, pady=2)

        self._iter_listbox = tk.Listbox(iter_frame, height=6, selectmode=tk.SINGLE,
                                         exportselection=False, width=18)
        self._iter_listbox.pack(side=tk.LEFT, padx=2, fill=tk.Y)
        for label in DEFAULT_ITER_ORDER:
            self._iter_listbox.insert(tk.END, label)

        btn_frame = ttk.Frame(iter_frame)
        btn_frame.pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="↑ 上移", width=8, command=self._iter_up).pack(pady=2)
        ttk.Button(btn_frame, text="↓ 下移", width=8, command=self._iter_down).pack(pady=2)

        ttk.Label(iter_frame, text="例: 图像模式在最上\n→ 同一图像模式下\n  峰值亮度先高再中再低",
                  foreground="gray", font=("", 8), justify=tk.LEFT).pack(side=tk.LEFT, padx=4)

        # 行颜色配置
        color_frame = ttk.LabelFrame(left, text="行颜色（按图像模式）", padding=6)
        color_frame.pack(fill=tk.X, pady=2)

        self._color_vars: dict[str, tk.StringVar] = {}
        for i, mode in enumerate(["标准", "影院", "鲜艳"]):
            ttk.Label(color_frame, text=f"{mode}:").grid(row=0, column=i * 2, padx=2)
            default_hex = DEFAULT_MODE_COLORS[mode]
            default_name = next((k for k, v in PRESET_COLORS.items() if v == default_hex), "浅灰")
            var = tk.StringVar(value=default_name)
            combo = ttk.Combobox(color_frame, textvariable=var, values=list(PRESET_COLORS.keys()),
                                 state="readonly", width=6)
            combo.grid(row=0, column=i * 2 + 1, padx=2)
            self._color_vars[mode] = var

        # ── 右侧预览区 ──
        right = ttk.Frame(top, padding=4)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        preview_lf = ttk.LabelFrame(right, text="预览", padding=4)
        preview_lf.pack(fill=tk.BOTH, expand=True)

        self._summary_label = ttk.Label(preview_lf, text="", font=("", 10), wraplength=400, justify=tk.LEFT)
        self._summary_label.pack(anchor=tk.W, pady=4)

        # 预览表格
        cols = ("编号", "图像模式", "峰值亮度", "对比度增强", "LD", "窗口", "HDR/SDR", "亮度")
        self._tree = ttk.Treeview(preview_lf, columns=cols, show="headings", height=15)
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=65, minwidth=50)
        self._tree.column("编号", width=55)
        self._tree.column("对比度增强", width=70)
        self._tree.column("LD", width=35)

        vsb = ttk.Scrollbar(preview_lf, orient=tk.VERTICAL, command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # ── 底部操作栏 ──
        bottom = ttk.Frame(root, padding=6)
        bottom.pack(fill=tk.X)

        self._generate_btn = ttk.Button(bottom, text="生成 Excel", command=self._generate)
        self._generate_btn.pack(side=tk.RIGHT, padx=4)

        self._status_var = tk.StringVar(value="就绪")
        ttk.Label(bottom, textvariable=self._status_var).pack(side=tk.LEFT, padx=4)

        # 初始化预览
        self._refresh_preview()

    # ──────────── 迭代顺序操作 ────────────

    def _iter_up(self):
        sel = self._iter_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        idx = sel[0]
        text = self._iter_listbox.get(idx)
        self._iter_listbox.delete(idx)
        self._iter_listbox.insert(idx - 1, text)
        self._iter_listbox.selection_set(idx - 1)
        self._refresh_preview()

    def _iter_down(self):
        sel = self._iter_listbox.curselection()
        if not sel or sel[0] == self._iter_listbox.size() - 1:
            return
        idx = sel[0]
        text = self._iter_listbox.get(idx)
        self._iter_listbox.delete(idx)
        self._iter_listbox.insert(idx + 1, text)
        self._iter_listbox.selection_set(idx + 1)
        self._refresh_preview()

    def _get_iteration_order(self) -> list[str]:
        """返回从最外到最内的因子 key 列表（列表框顺序）。"""
        return [ITER_LABEL_TO_KEY[self._iter_listbox.get(i)]
                for i in range(self._iter_listbox.size())]

    # ──────────── 数据收集 ────────────

    def _get_windows(self) -> list[str]:
        wins = []
        start_val = self._win_start.get().strip()
        end_val = self._win_end.get().strip()
        step_val = self._win_step.get().strip()

        if start_val and end_val:
            try:
                start = int(start_val)
                end = int(end_val)
                step = int(step_val) if step_val else None
                if step is not None and step > 0 and start <= end:
                    wins = [f"{v}%" for v in range(start, end + 1, step)]
                elif start <= end:
                    wins = [f"{start}%", f"{end}%"] if start != end else [f"{start}%"]
            except ValueError:
                pass

        extra = self._win_extra.get().strip()
        if extra:
            for v in extra.replace("，", ",").split(","):
                v = v.strip().rstrip("%")
                label = f"{v}%"
                if label not in wins:
                    wins.append(label)

        def sort_key(s):
            try:
                return int(s.rstrip("%"))
            except ValueError:
                return 999
        wins.sort(key=sort_key)
        return wins

    @staticmethod
    def _parse_int_range(start_val: str, end_val: str, step_val: str, extra_val: str) -> list[int]:
        """从起始/结束/间隔/额外 解析出整数列表。"""
        result = []
        if start_val and end_val:
            try:
                start = int(start_val)
                end = int(end_val)
                step = int(step_val) if step_val else None
                if step is not None and step > 0:
                    result.extend(range(start, end + 1, step))
                elif start <= end:
                    result.extend([start, end] if start != end else [start])
            except ValueError:
                pass
        if extra_val:
            for v in extra_val.replace("，", ",").split(","):
                v = v.strip()
                if v:
                    try:
                        n = int(v)
                        if n not in result:
                            result.append(n)
                    except ValueError:
                        pass
        result.sort()
        return result

    def _get_brightness(self) -> dict[str, list[int]]:
        """返回 {"HDR": [...], "SDR": [...]} 格式的亮度值。"""
        result = {}
        for label, vars_ in self._bright_vars.items():
            vals = self._parse_int_range(
                vars_["start"].get().strip(),
                vars_["end"].get().strip(),
                vars_["step"].get().strip(),
                vars_["extra"].get().strip(),
            )
            result[label] = vals
        return result

    def _collect_factors(self) -> dict:
        bright_map = self._get_brightness()
        # 亮度取 HDR/SDR 的并集（用于预览公式和因子水平表）
        all_bright = sorted(set(bright_map.get("HDR", []) + bright_map.get("SDR", [])))
        return {
            "picture_mode": self._factor_frames["picture_mode"].get_values(),
            "peak_brightness": self._factor_frames["peak_brightness"].get_values(),
            "contrast_enhance": self._factor_frames["contrast_enhance"].get_values(),
            "local_dimming": self._factor_frames["local_dimming"].get_values(),
            "hdr_sdr": self._factor_frames["hdr_sdr"].get_values(),
            "brightness": all_bright,
            "window": self._get_windows(),
        }

    def _generate_test_cases(self) -> list[dict]:
        factors = self._collect_factors()
        bright_map = self._get_brightness()
        order_outer_to_inner = self._get_iteration_order()

        active_order = [k for k in order_outer_to_inner if factors.get(k)]
        if not active_order:
            return []

        test_cases = []

        def _recurse(depth: int, current: dict):
            if depth == len(active_order):
                test_cases.append({
                    "picture_mode": current.get("picture_mode", ""),
                    "peak_brightness": current.get("peak_brightness", ""),
                    "contrast_enhance": current.get("contrast_enhance", ""),
                    "local_dimming": current.get("local_dimming", ""),
                    "window": current.get("window", ""),
                    "hdr_sdr": current.get("hdr_sdr", ""),
                    "brightness": current.get("brightness", ""),
                })
                return
            key = active_order[depth]
            # 白块亮度根据当前 HDR/SDR 取对应列表
            if key == "brightness":
                hdr_val = current.get("hdr_sdr", "")
                vals = bright_map.get(hdr_val, [])
                if not vals:
                    return
                for val in vals:
                    current[key] = val
                    _recurse(depth + 1, current)
            else:
                for val in factors[key]:
                    current[key] = val
                    _recurse(depth + 1, current)
            current.pop(key, None)

        _recurse(0, {})
        return test_cases

    # ──────────── 预览 ────────────

    def _refresh_preview(self):
        test_cases = self._generate_test_cases()
        total = len(test_cases)

        # Summary
        factors = self._collect_factors()
        label_map = {
            "window": "窗口", "brightness": "亮度", "hdr_sdr": "HDR/SDR",
            "peak_brightness": "峰值亮度", "contrast_enhance": "对比度增强",
            "local_dimming": "LD", "picture_mode": "图像模式",
        }
        # 按迭代顺序从外到内显示公式
        parts = []
        for key in self._get_iteration_order():
            vals = factors.get(key, [])
            if vals:
                parts.append(f"{len(vals)}({label_map[key]})")
        formula = " × ".join(parts) if parts else "0"
        self._summary_label.config(text=f"总计: {total} 组\n公式: {formula}\n迭代: {' → '.join(label_map[k] for k in self._get_iteration_order() if factors.get(k))} (外→内)")

        # Tree preview (first 50)
        self._tree.delete(*self._tree.get_children())
        for i, tc in enumerate(test_cases[:50]):
            self._tree.insert("", tk.END, values=(
                f"T{i+1:03d}", tc["picture_mode"], tc["peak_brightness"],
                tc["contrast_enhance"], tc["local_dimming"], tc["window"],
                tc["hdr_sdr"], tc["brightness"],
            ))
        if total > 50:
            self._tree.insert("", tk.END, values=(f"... 共 {total} 组", "", "", "", "", "", "", ""))

    # ──────────── Excel 生成 ────────────

    def _generate(self):
        test_cases = self._generate_test_cases()
        if not test_cases:
            messagebox.showwarning("提示", "无测试用例可生成，请检查参数设置。")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            title="保存测试用例",
            initialfile="峰值亮度测试用例.xlsx",
        )
        if not path:
            return

        self._status_var.set("正在生成...")
        self.root.update_idletasks()

        try:
            self._write_excel(path, test_cases)
            self._status_var.set(f"已完成: {path}")
            messagebox.showinfo("完成", f"已生成 {len(test_cases)} 组测试用例\n保存至: {path}")
        except Exception as e:
            self._status_var.set("生成失败")
            messagebox.showerror("错误", str(e))

    def _write_excel(self, path: str, test_cases: list[dict]):
        bright_map = self._get_brightness()
        wb = Workbook()

        # ── Sheet 1: 测试用例 ──
        ws = wb.active
        ws.title = "测试用例"

        # Row 1: section headers
        ws.merge_cells("A1:E1")
        ws.merge_cells("F1:H1")
        ws.merge_cells("I1:K1")

        for cell_ref, label, fill in [
            ("A1", "电视端设置", BLUE_FILL),
            ("F1", "MURIDEO 8K SEVEN", ORANGE_FILL),
            ("I1", "CA-410 测量数据", BLUE_FILL),
        ]:
            c = ws[cell_ref]
            c.value = label
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.fill = fill
            c.border = THIN_BORDER

        # Fill merged region styling
        for col_idx in range(1, 6):
            c = ws.cell(row=1, column=col_idx)
            c.fill = BLUE_FILL
            c.border = THIN_BORDER
        for col_idx in range(6, 9):
            c = ws.cell(row=1, column=col_idx)
            c.fill = ORANGE_FILL
            c.border = THIN_BORDER
        for col_idx in range(9, 12):
            c = ws.cell(row=1, column=col_idx)
            c.fill = BLUE_FILL
            c.border = THIN_BORDER
        ws.cell(row=1, column=12).fill = BLUE_FILL
        ws.cell(row=1, column=12).border = THIN_BORDER

        # Row 2: column headers
        headers = ["编号", "图像模式", "峰值亮度", "对比度增强", "Local Dimming",
                    "小窗口大小", "HDR/SDR", "白块亮度(nit)",
                    "Lv (cd/m²)", "x", "y", "备注"]
        col_widths = [14, 12, 12, 14, 17, 25, 11, 17, 19, 5, 5, 59]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER
            c.fill = ORANGE_FILL if 6 <= i <= 8 else BLUE_FILL
            ws.column_dimensions[get_column_letter(i)].width = w

        # Build color map from user selections
        color_map = {}
        for mode, var in self._color_vars.items():
            color_name = var.get()
            color_map[mode] = PatternFill(
                start_color=PRESET_COLORS.get(color_name, "F2F2F2"),
                end_color=PRESET_COLORS.get(color_name, "F2F2F2"),
                fill_type="solid",
            )

        # Data rows
        for i, tc in enumerate(test_cases):
            row = i + 3
            note = (f"测试{tc['picture_mode']}模式{tc['window']}窗口"
                    f"峰值亮度{tc['peak_brightness']}的亮度输出；"
                    f"{tc['hdr_sdr']}；白块亮度{tc['brightness']}nit")
            values = [
                f"T{i+1:03d}", tc["picture_mode"], tc["peak_brightness"],
                tc["contrast_enhance"], tc["local_dimming"],
                tc["window"], tc["hdr_sdr"], tc["brightness"],
                None, None, None, note,
            ]
            row_fill = color_map.get(tc["picture_mode"], PatternFill())
            for col_idx, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.alignment = CENTER
                c.border = THIN_BORDER
                c.fill = row_fill

        ws.freeze_panes = "A3"

        # ── Sheet 2: 测试计划 ──
        ws2 = wb.create_sheet("测试计划")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 80

        factors = self._collect_factors()
        label_map = {
            "window": "窗口大小", "brightness": "白块亮度", "hdr_sdr": "HDR/SDR",
            "peak_brightness": "峰值亮度", "contrast_enhance": "对比度增强",
            "local_dimming": "Local Dimming", "picture_mode": "图像模式",
        }
        iter_labels = [label_map[k] for k in self._get_iteration_order() if factors.get(k)]

        plan_items = [
            ("测试目标", "探究图像模式、峰值亮度、白块亮度、小窗口大小、HDR/SDR、Local Dimming对峰值亮度的影响规律"),
            ("图像模式", "、".join(factors["picture_mode"]) or "无"),
            ("峰值亮度", "、".join(factors["peak_brightness"]) or "无"),
            ("白块亮度", "；".join(f"{k}: {', '.join(str(b) for b in v)} nit" for k, v in bright_map.items() if v) or "无"),
            ("小窗口大小", "、".join(factors["window"]) or "无"),
            ("HDR/SDR", "、".join(factors["hdr_sdr"]) or "无"),
            ("Local Dimming", "、".join(factors["local_dimming"]) or "无"),
            ("对比度增强", "、".join(factors["contrast_enhance"])),
            ("迭代顺序", " → ".join(iter_labels) + "（最外→最内，先改变最内层）"),
            ("信号源", "MURIDEO 8K SEVEN GENERATOR"),
            ("测量设备", "CA-410 色彩分析仪"),
            ("总计", f"{len(test_cases)}组"),
            ("每组测量次数", "1次"),
            ("操作要点1", "HDR时MURIDEO输出HDR10/PQ EOTF/BT.2020格式；SDR时输出BT.709/SDR格式"),
            ("操作要点2", "每次切换参数后等待3秒，确保屏幕稳定"),
            ("操作要点3", "保持环境光恒定（暗室或固定照度）"),
            ("操作要点4", "CA-410探头对准MURIDEO输出白块中心位置"),
            ("操作要点5", "关注色度变化：峰值亮度档位可能改变x/y值，不仅看Lv"),
        ]
        legend_parts = []
        for mode, var in self._color_vars.items():
            legend_parts.append(f"{var.get()}={mode}")
        plan_items.append(("行颜色说明", "；".join(legend_parts)))

        for r, (k, v) in enumerate(plan_items, 1):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
            ws2.cell(row=r, column=2, value=v)

        # ── Sheet 3: 因子水平 ──
        ws3 = wb.create_sheet("因子水平")
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 35
        ws3.column_dimensions["C"].width = 15
        ws3.column_dimensions["D"].width = 40
        ws3.column_dimensions["E"].width = 15

        factor_headers = ["因子", "水平", "类型", "说明", "来源"]
        for i, h in enumerate(factor_headers, 1):
            c = ws3.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.fill = BLUE_FILL
            c.border = THIN_BORDER

        factor_data = [
            ("A: 图像模式", "、".join(factors["picture_mode"]), "离散", "显示预设模式", "电视端"),
            ("B: 峰值亮度", "、".join(factors["peak_brightness"]), "有序离散", "峰值亮度档位", "电视端"),
            ("C: 对比度增强", "、".join(factors["contrast_enhance"]), "有序离散", "对比度增强档位", "电视端"),
            ("D: Local Dimming", "、".join(factors["local_dimming"]), "有序离散", "Local Dimming档位", "电视端"),
            ("E: 小窗口大小", f"共{len(factors['window'])}级", "连续", "信号源输出测试图案窗口大小", "MURIDEO"),
            ("F: HDR/SDR", "、".join(factors["hdr_sdr"]), "离散", "HDR10/PQ/BT.2020 或 SDR/BT.709", "MURIDEO"),
            ("G: 白块亮度", "；".join(f"{k}: {', '.join(str(b) for b in v)}" for k, v in bright_map.items() if v), "连续", "HDR/SDR分别设置白块亮度", "MURIDEO"),
        ]
        for r, row_data in enumerate(factor_data, 2):
            for c_idx, val in enumerate(row_data, 1):
                c = ws3.cell(row=r, column=c_idx, value=val)
                c.alignment = CENTER
                c.border = THIN_BORDER

        # ── Sheet 4: MURIDEO设置参考 ──
        ws4 = wb.create_sheet("MURIDEO设置参考")
        ws4.column_dimensions["A"].width = 20
        ws4.column_dimensions["B"].width = 14
        ws4.column_dimensions["C"].width = 14
        ws4.column_dimensions["D"].width = 25
        ws4.column_dimensions["E"].width = 18

        ref_headers = ["测试编号段", "图像模式", "峰值亮度", "小窗口大小", "白块亮度(nit)"]
        for i, h in enumerate(ref_headers, 1):
            c = ws4.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.fill = ORANGE_FILL
            c.border = THIN_BORDER

        from collections import OrderedDict
        groups = OrderedDict()
        for i, tc in enumerate(test_cases):
            key = (tc["picture_mode"], tc["peak_brightness"], tc["hdr_sdr"], tc["brightness"])
            if key not in groups:
                groups[key] = []
            groups[key].append(i)

        ref_row = 2
        for key, indices in groups.items():
            mode, peak, hdr, bright = key
            start_id = indices[0] + 1
            end_id = indices[-1] + 1
            ws4.cell(row=ref_row, column=1, value=f"T{start_id:03d}-T{end_id:03d}").border = THIN_BORDER
            ws4.cell(row=ref_row, column=2, value=mode).border = THIN_BORDER
            ws4.cell(row=ref_row, column=3, value=peak).border = THIN_BORDER
            ws4.cell(row=ref_row, column=4, value="、".join(factors["window"])).border = THIN_BORDER
            ws4.cell(row=ref_row, column=5, value=bright).border = THIN_BORDER
            for c in range(1, 6):
                ws4.cell(row=ref_row, column=c).alignment = CENTER
            ref_row += 1

        ref_row += 1
        common = [
            ("通用设置", ""),
            ("分辨率", "3840×2160 — 4K匹配电视"),
            ("色深", "10bit — 避免色带影响测量"),
            ("HDR格式", "HDR10 / SDR"),
            ("EOTF", "PQ (ST.2084) / BT.1886"),
            ("色域", "BT.2020 / BT.709"),
        ]
        for label, val in common:
            c1 = ws4.cell(row=ref_row, column=1, value=label)
            c2 = ws4.cell(row=ref_row, column=2, value=val)
            if label == "通用设置":
                c1.font = Font(bold=True, size=11)
            ref_row += 1

        wb.save(path)


def main():
    root = tk.Tk()
    GeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
