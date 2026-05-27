"""Load test cases from the Boost peak brightness Excel file."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class TestCase:
    """A single test case from the Excel file or profile."""
    row: int                    # Excel row number (1-indexed)
    test_id: str                # 编号, e.g. "T001"
    image_mode: str             # 图像模式
    peak_brightness: str        # 峰值亮度
    backlight_value: str        # 对比度增强
    local_dimming: str          # Local Dimming
    window_size: float          # 小窗口大小 (%), e.g. 10, 50, 100
    hdr_sdr: str                # HDR/SDR
    window_brightness: float    # IRE Level (0-255)
    note: str                   # 备注
    # Extended Murideo parameters (optional, used by profile-based tests)
    timing: int | None = None           # 信号格式 ID, e.g. 34=3840x2160@60Hz
    color_space: int | None = None      # 0=RGB(0-255), 1=RGB(16-235), 2=YC444, 3=YC422, 4=YC420
    color_depth: int | None = None      # 0=8bit, 1=10bit, 2=12bit, 3=16bit
    pattern: int | None = None          # 图案 ID, e.g. 26=Window, 11=White


def load_test_cases(filepath: str | Path) -> list[TestCase]:
    """Load all test cases from the Excel file.

    Expected sheet name: "测试用例"
    Expected column layout (row 1 = category headers, row 2 = column headers):
        Col 1: 编号
        Col 2: 图像模式
        Col 3: 峰值亮度
        Col 4: 对比度增强
        Col 5: Local Dimming
        Col 6: 小窗口大小 (e.g. "10%", "50%", "100%")
        Col 7: HDR/SDR
        Col 8: 白块亮度(nit)
        Col 9-11: measurement data (Lv, x, y) — filled later
        Col 12: 备注
    """
    wb = load_workbook(str(filepath), data_only=True)
    ws = wb['测试用例']

    cases = []
    for row in range(3, ws.max_row + 1):
        test_id = ws.cell(row=row, column=1).value
        if not test_id:
            continue

        image_mode = _str(ws.cell(row=row, column=2).value)
        peak_brightness = _str(ws.cell(row=row, column=3).value)
        backlight_value = _str(ws.cell(row=row, column=4).value)
        local_dimming = _str(ws.cell(row=row, column=5).value)
        window_size = _parse_percent(ws.cell(row=row, column=6).value)
        hdr_sdr = _str(ws.cell(row=row, column=7).value)
        window_brightness = _num(ws.cell(row=row, column=8).value, default=0)
        note = _str(ws.cell(row=row, column=12).value)

        cases.append(TestCase(
            row=row,
            test_id=test_id,
            image_mode=image_mode,
            peak_brightness=peak_brightness,
            backlight_value=backlight_value,
            local_dimming=local_dimming,
            window_size=window_size,
            hdr_sdr=hdr_sdr,
            window_brightness=window_brightness,
            note=note,
        ))

    wb.close()
    return cases


def _str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _num(val, default: float = 0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _parse_percent(val) -> float:
    """Parse '10%' -> 10, '50%' -> 50, '100' -> 100."""
    if val is None:
        return 0
    s = str(val).strip()
    if s.endswith('%'):
        s = s[:-1].strip()
    try:
        return float(s)
    except ValueError:
        return 0
