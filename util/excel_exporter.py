"""Export measurement data to the Boost peak brightness test case Excel file.

Matched records are written back to the matching row in the "测试用例" sheet.
Unmatched records are appended to the "额外数据" sheet.
"""
import tkinter.filedialog as fd
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Column mapping in the Excel test case sheet (1-indexed)
COL_MAP = {
    '编号': 1,
    '图像模式': 2,
    '峰值亮度': 3,
    '当前背光值': 4,
    'Local Dimming': 5,
    '小窗口大小': 6,
    'HDR/SDR': 7,
    '白块亮度(nit)': 8,
    'Lv (cd/m²)': 9,
    'x': 10,
    'y': 11,
    '备注': 12,
}

# Keys used for matching a measurement to a test case row
MATCH_KEYS = ['图像模式', '峰值亮度', '当前背光值', 'Local Dimming', '小窗口大小', 'HDR/SDR', '白块亮度(nit)']

EXTRA_SHEET = '额外数据'


def _norm(val) -> str:
    """Normalize a value for comparison: strip, remove % suffix, lowercase."""
    if val is None:
        return ''
    s = str(val).strip()
    # Normalize percentage: "10%" -> "10", "10.0%" -> "10"
    if s.endswith('%'):
        s = s[:-1].strip()
    # Normalize numeric: "100.0" -> "100"
    try:
        f = float(s)
        if f == int(f):
            s = str(int(f))
    except ValueError:
        pass
    return s.lower()


def _match_row(ws, row: int, record: dict) -> bool:
    """Check if a test case row matches the measurement record."""
    for key in MATCH_KEYS:
        cell_val = ws.cell(row=row, column=COL_MAP[key]).value
        rec_val = record.get(key, '')
        if _norm(cell_val) != _norm(rec_val):
            return False
    return True


def _find_match_row(ws, record: dict, max_row: int) -> int | None:
    """Find the first test case row matching the record. Returns row number or None."""
    for row in range(3, max_row + 1):
        if ws.cell(row=row, column=1).value is None:
            continue
        if _match_row(ws, row, record):
            return row
    return None


def _write_to_row(ws, row: int, record: dict) -> None:
    """Write Lv, x, y to a test case row."""
    center = Alignment(horizontal='center', vertical='center')
    if record.get('Lv (cd/m²)'):
        c = ws.cell(row=row, column=COL_MAP['Lv (cd/m²)'], value=float(record['Lv (cd/m²)']))
        c.alignment = center
        c.number_format = '0.00'
    if record.get('x'):
        c = ws.cell(row=row, column=COL_MAP['x'], value=float(record['x']))
        c.alignment = center
        c.number_format = '0.0000'
    if record.get('y'):
        c = ws.cell(row=row, column=COL_MAP['y'], value=float(record['y']))
        c.alignment = center
        c.number_format = '0.0000'


def _ensure_extra_sheet(wb) -> None:
    """Create the extra data sheet if it doesn't exist."""
    if EXTRA_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(EXTRA_SHEET)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = list(COL_MAP.keys())
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border


def _append_extra(ws, record: dict) -> None:
    """Append an unmatched record to the extra data sheet."""
    row = ws.max_row + 1
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for key, col in COL_MAP.items():
        val = record.get(key, '')
        # Convert numeric strings to numbers for proper formatting
        if key in ('Lv (cd/m²)',) and val:
            val = float(val)
        elif key in ('x', 'y') and val:
            val = float(val)
        elif key in ('当前背光值', '白块亮度(nit)') and val:
            try:
                val = float(val)
                if val == int(val):
                    val = int(val)
            except ValueError:
                pass
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = center
        c.border = thin_border
        if key == 'Lv (cd/m²)':
            c.number_format = '0.00'
        elif key in ('x', 'y'):
            c.number_format = '0.0000'


def export_to_excel(data: list[dict], filepath: str) -> tuple[int, int]:
    """Write measurement data to the test case Excel file.

    Returns (matched_count, unmatched_count).
    """
    wb = load_workbook(filepath)
    ws = wb['测试用例']
    max_row = ws.max_row

    matched = 0
    unmatched = 0

    for record in data:
        row = _find_match_row(ws, record, max_row)
        if row is not None:
            _write_to_row(ws, row, record)
            matched += 1
        else:
            _ensure_extra_sheet(wb)
            ws_extra = wb[EXTRA_SHEET]
            _append_extra(ws_extra, record)
            unmatched += 1

    wb.save(filepath)
    return matched, unmatched


def prompt_excel_path(parent) -> str | None:
    """Ask user to select or create the test case Excel file."""
    filepath = fd.asksaveasfilename(
        parent=parent,
        defaultextension='.xlsx',
        filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')],
        initialfile='Boost峰值亮度测试用例.xlsx',
    )
    return filepath if filepath else None
