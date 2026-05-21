"""Generate Excel test cases for peak brightness study.

Full cross: Image mode × Peak brightness × Brightness × Window size
Iteration order (outermost→innermost): 图像模式 → 峰值亮度 → 亮度 → 窗口
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

ws = wb.active
ws.title = "测试用例"

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
murideo_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Section headers (row 1)
ws.merge_cells("A1:E1")
ws.merge_cells("F1:H1")
ws.merge_cells("I1:K1")
ws.merge_cells("L1:L1")

section_labels = [
    ("A1", "电视端设置", header_fill),
    ("F1", "MURIDEO 8K SEVEN\n信号源设置", murideo_fill),
    ("I1", "CA-410 测量数据", header_fill),
    ("L1", "", header_fill),
]
for cell_ref, label, fill in section_labels:
    c = ws[cell_ref]
    c.value = label
    c.font = header_font_white
    c.fill = fill
    c.alignment = center
    c.border = thin_border

# Row 2: column headers
headers = [
    "编号", "图像模式", "峰值亮度", "当前背光值", "Local Dimming",
    "小窗口大小", "HDR/SDR", "白块亮度(nit)",
    "Lv (cd/m²)", "x", "y",
    "备注"
]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = header_font_white
    if i in (6, 7, 8):
        c.fill = murideo_fill
    else:
        c.fill = header_fill
    c.alignment = center
    c.border = thin_border

ws.row_dimensions[1].height = 45
ws.row_dimensions[2].height = 45

# ----- Parameters -----
IMAGE_MODES = ["鲜艳", "影院", "标准"]
PEAK_LEVELS = ["高", "中", "低"]
BRIGHTNESSES = [128, 144, 167, 200]
WINDOW_SIZES = list(range(0, 101, 5))  # 0%, 5%, 10%, ..., 100%

# ----- Generate test cases -----
# Order: 图像模式(outermost) → 峰值亮度 → 亮度 → 窗口(innermost)
test_cases = []
idx = 0

def _note(mode, peak, bright, win):
    if peak == "高":
        return f"测试{mode}模式{win}%窗口峰值亮度{peak}的亮度输出；白块亮度{bright}nit"
    elif peak == "中":
        return f"测试{mode}模式{win}%窗口峰值亮度{peak}的亮度输出；白块亮度{bright}nit"
    else:
        return f"测试{mode}模式{win}%窗口峰值亮度{peak}的亮度输出；白块亮度{bright}nit"

for mode in IMAGE_MODES:
    for peak in PEAK_LEVELS:
        for bright in BRIGHTNESSES:
            for win in WINDOW_SIZES:
                idx += 1
                test_cases.append({
                    "编号": f"T{idx:03d}",
                    "图像模式": mode,
                    "峰值亮度": peak,
                    "当前背光值": 100,
                    "Local Dimming": "强",
                    "窗口大小": f"{win}%",
                    "hdr_sdr": "HDR",
                    "白块亮度": bright,
                    "note": _note(mode, peak, bright, win),
                })

# Row fill: color by image mode for easy visual scanning
mode_fills = {
    "鲜艳": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "影院": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "标准": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}

for tc in test_cases:
    row = int(tc["编号"][1:]) + 2
    ws.row_dimensions[row].height = 22
    values = [
        tc["编号"], tc["图像模式"], tc["峰值亮度"],
        tc["当前背光值"], tc["Local Dimming"],
        tc["窗口大小"], tc["hdr_sdr"], tc["白块亮度"],
        "", "", "",
        tc["note"]
    ]
    row_fill = mode_fills[tc["图像模式"]]
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.alignment = center
        c.border = thin_border
        c.fill = row_fill

ws.freeze_panes = "A3"


# ===== Sheet 2: Test plan =====
ws2 = wb.create_sheet("测试计划")

plan_items = [
    ("测试目标", "探究图像模式、峰值亮度档位、白块亮度、小窗口大小对峰值亮度的影响规律"),
    ("图像模式", "鲜艳、影院、标准"),
    ("峰值亮度", "高、中、低"),
    ("白块亮度", "128、144、167、200 nit"),
    ("小窗口大小", "0%-100%（5%间隔，共21级）"),
    ("迭代顺序", "先改变窗口大小 → 再改变亮度 → 再改变峰值亮度 → 最后改变图像模式"),
    ("信号源", "MURIDEO 8K SEVEN GENERATOR"),
    ("测量设备", "CA-410 色彩分析仪"),
    ("总计", f"3(图像模式) × 3(峰值亮度) × 4(亮度) × 21(窗口) = {len(test_cases)}组"),
    ("固定条件", "当前背光值=100; Local Dimming=强; HDR"),
    ("每组测量次数", "1次"),
    ("操作要点1", "HDR时MURIDEO输出HDR10/PQ EOTF/BT.2020格式"),
    ("操作要点2", "每次切换参数后等待3秒，确保屏幕稳定"),
    ("操作要点3", "保持环境光恒定（暗室或固定照度）"),
    ("操作要点4", "CA-410探头对准MURIDEO输出白块中心位置"),
    ("操作要点5", "关注色度变化：峰值亮度档位可能改变x/y值，不仅看Lv"),
    ("行颜色说明", "橙色=鲜艳, 蓝色=影院, 绿色=标准"),
]

for i, (k, v) in enumerate(plan_items, 1):
    ck = ws2.cell(row=i, column=1, value=k)
    ck.font = Font(bold=True, size=11)
    cv = ws2.cell(row=i, column=2, value=v)
    cv.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[i].height = 30 if "\n" in str(v) else 20


# ===== Sheet 3: Factor levels =====
ws3 = wb.create_sheet("因子水平")

ref_headers = ["因子", "水平", "类型", "说明", "来源"]
for i, h in enumerate(ref_headers, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

ref_data = [
    ("A: 图像模式", "鲜艳、影院、标准", "离散", "显示预设模式，不同模式对峰值亮度输出可能不同", "电视端"),
    ("B: 峰值亮度", "高、中、低", "有序离散", "峰值亮度档位", "电视端"),
    ("C: 当前背光值", "100 (固定)", "连续", "背光PWM/亮度设置值，本次测试固定为100", "电视端"),
    ("D: Local Dimming", "强 (固定)", "有序离散", "本次测试固定为强", "电视端"),
    ("E: 小窗口大小", "0%-100%(5%间隔)=21级", "连续", "信号源输出测试图案窗口大小", "MURIDEO"),
    ("F: HDR/SDR", "HDR (固定)", "离散", "HDR10/PQ/BT.2020", "MURIDEO"),
    ("G: 白块亮度", "128/144/167/200 nit", "连续", "信号源输出白块亮度", "MURIDEO"),
]
for r, (factor, levels, ftype, desc, source) in enumerate(ref_data, 2):
    for c_idx, val in enumerate([factor, levels, ftype, desc, source], 1):
        c = ws3.cell(row=r, column=c_idx, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if source == "MURIDEO":
            c.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")


# ===== Sheet 4: MURIDEO setup reference =====
ws4 = wb.create_sheet("MURIDEO设置参考")

setup_headers = ["测试编号段", "图像模式", "峰值亮度", "小窗口大小", "白块亮度(nit)"]
for i, h in enumerate(setup_headers, 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = murideo_fill
    c.alignment = center
    c.border = thin_border

setup_data = []
idx = 0
for mode in IMAGE_MODES:
    for peak in PEAK_LEVELS:
        for bright in BRIGHTNESSES:
            start = idx + 1
            end = idx + len(WINDOW_SIZES)
            idx = end
            setup_data.append((
                f"T{start:03d}-T{end:03d}",
                mode, peak,
                "0%-100%(5%间隔)",
                str(bright)
            ))

for r, vals in enumerate(setup_data, 2):
    for c_idx, val in enumerate(vals, 1):
        c = ws4.cell(row=r, column=c_idx, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")

# MURIDEO通用设置
setup_data2 = [
    ("", "", "", "", ""),
    ("通用设置", "", "", "", ""),
    ("分辨率", "3840×2160", "", "", "4K匹配电视"),
    ("色深", "10bit", "", "", "避免色带影响测量"),
    ("HDR格式", "HDR10", "", "", "HDR模式"),
    ("EOTF", "PQ (ST.2084)", "", "", "HDR10标准EOTF"),
    ("色域", "BT.2020", "", "", "匹配电视广色域模式"),
]
for vals in setup_data2:
    r = ws4.max_row + 1
    for c_idx, val in enumerate(vals, 1):
        c = ws4.cell(row=r, column=c_idx, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if val == "通用设置":
            c.font = Font(bold=True, size=11)


# Auto-fit column widths for all sheets
def autofit_columns(sheet):
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val = str(cell.value)
                line_max = max(len(line) for line in val.split('\n'))
                cjk_len = sum(1 for ch in val if '一' <= ch <= '鿿' or '　' <= ch <= '〿')
                adjusted = line_max + cjk_len
                if adjusted > max_len:
                    max_len = adjusted
        if max_len > 0:
            sheet.column_dimensions[col_letter].width = max_len + 4

for sheet in wb.worksheets:
    autofit_columns(sheet)

# Save
output = "/home/liaoxiaolan/claudeCodeProject/read410/峰值亮度测试用例.xlsx"
wb.save(output)
print(f"Saved to: {output}")
print(f"Total test cases: {len(test_cases)}")
