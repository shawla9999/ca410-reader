"""Generate Excel test cases for Boost peak brightness study.

Full cross: Image mode × Window size × Peak brightness × White block brightness
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

ws = wb.active
ws.title = "测试用例"

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
murideo_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Section headers (row 1)
# TV: A-E (5 cols), MURIDEO: F-H (3 cols), CA-410: I-K (3 cols), Note: L (1 col)
ws.merge_cells("A1:E1")
ws.merge_cells("F1:H1")
ws.merge_cells("I1:K1")
ws.merge_cells("L1:L1")

section_labels = [
    ("A1", "电视端设置", header_fill),
    ("F1", "MURIDEO 8K SEVEN\n信号源设置", murideo_fill),
    ("I1", "CA-410 测量数据", header_fill),
    ("L1", "", header_fill),
]
for cell_ref, label, fill in section_labels:
    c = ws[cell_ref]
    c.value = label
    c.font = header_font_white
    c.fill = fill
    c.alignment = center
    c.border = thin_border

# Row 2: column headers
headers = [
    # TV (A-E)
    "编号", "图像模式", "峰值亮度", "当前背光值", "Local Dimming",
    # MURIDEO (F-H)
    "小窗口大小", "HDR/SDR", "白块亮度(nit)",
    # CA-410 (I-K)
    "Lv (cd/m²)", "x", "y",
    # Note
    "备注"
]

col_widths = [7, 10, 10, 12, 14,
              12, 10, 14,
              14, 10, 10,
              20]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = header_font_white
    if i in (6, 7, 8):
        c.fill = murideo_fill
    else:
        c.fill = header_fill
    c.alignment = center
    c.border = thin_border

ws.row_dimensions[1].height = 45
ws.row_dimensions[2].height = 45

# ----- Generate full cross test cases -----
IMAGE_MODES = ["鲜艳", "标准", "影院"]
WINDOW_SIZES = [10, 50, 100]
PEAK_LEVELS = ["关", "弱", "中", "强"]
BRIGHTNESSES = [100, 200, 600]

test_cases = []
idx = 0

def _hdr_note(mode, win, peak, bright):
    """Generate purpose note for HDR cases."""
    parts = []
    if peak == "关":
        parts.append(f"{mode}模式{win}%窗口boost关闭基线")
    else:
        parts.append(f"测试{mode}模式{win}%窗口boost{peak}的亮度提升")
    if bright == 100:
        parts.append("验证低亮度片源下boost是否生效")
    elif bright == 200:
        parts.append("验证中低亮度片源下boost表现")
    else:
        parts.append("正常亮度片源下的boost表现")
    return "；".join(parts)


def _sdr_note(mode, win, peak, bright):
    """Generate purpose note for SDR cases."""
    if peak == "关":
        return f"{mode}模式SDR信号{win}%窗口boost关闭基线"
    return f"测试{mode}模式SDR信号下boost{peak}是否生效，与HDR同条件对比"


def _ld_note(mode, win, ld):
    """Generate purpose note for LD×Boost cases."""
    return f"测试{mode}模式{win}%窗口boost强+LD{ld}的亮度输出，与LD关对比分析LD对boost的影响"


# HDR cases: 3 modes × 3 windows × 4 peaks × 3 brightnesses = 108
for mode in IMAGE_MODES:
    for win in WINDOW_SIZES:
        for peak in PEAK_LEVELS:
            for bright in BRIGHTNESSES:
                idx += 1
                test_cases.append({
                    "编号": f"T{idx:03d}",
                    "图像模式": mode,
                    "峰值亮度": peak,
                    "当前背光值": 100,
                    "Local Dimming": "关",
                    "窗口大小": f"{win}%",
                    "hdr_sdr": "HDR",
                    "白块亮度": bright,
                    "note": _hdr_note(mode, win, peak, bright),
                })

# SDR cases: 3 modes × 2 windows × 4 peaks × 1 brightness = 24
SDR_WINDOW_SIZES = [10, 50]
SDR_BRIGHTNESSES = [100]

for mode in IMAGE_MODES:
    for win in SDR_WINDOW_SIZES:
        for peak in PEAK_LEVELS:
            for bright in SDR_BRIGHTNESSES:
                idx += 1
                test_cases.append({
                    "编号": f"T{idx:03d}",
                    "图像模式": mode,
                    "峰值亮度": peak,
                    "当前背光值": 100,
                    "Local Dimming": "关",
                    "窗口大小": f"{win}%",
                    "hdr_sdr": "SDR",
                    "白块亮度": bright,
                    "note": _sdr_note(mode, win, peak, bright),
                })

# LD × boost强 interaction: 3 modes × 2 windows × peak=强 × LD=弱/中/强 = 18
# LD=关+peak=强 already covered in HDR group, so only add 弱/中/强
for mode in IMAGE_MODES:
    for win in [10, 50]:
        for ld in ["弱", "中", "强"]:
            idx += 1
            test_cases.append({
                "编号": f"T{idx:03d}",
                "图像模式": mode,
                "峰值亮度": "强",
                "当前背光值": 100,
                "Local Dimming": ld,
                "窗口大小": f"{win}%",
                "hdr_sdr": "HDR",
                "白块亮度": 100,
                "note": _ld_note(mode, win, ld),
            })

# Row fill: color by (mode, hdr_sdr) for easy visual scanning
mode_hdr_fills = {
    ("鲜艳", "HDR"): PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ("标准", "HDR"): PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    ("影院", "HDR"): PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    ("鲜艳", "SDR"): PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    ("标准", "SDR"): PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
    ("影院", "SDR"): PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    ("鲜艳", "LD"): PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    ("标准", "LD"): PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
    ("影院", "LD"): PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),
}

for tc in test_cases:
    row = int(tc["编号"][1:]) + 2
    ws.row_dimensions[row].height = 22
    values = [
        tc["编号"], tc["图像模式"], tc["峰值亮度"],
        tc["当前背光值"], tc["Local Dimming"],
        tc["窗口大小"], tc["hdr_sdr"], tc["白块亮度"],
        "", "", "",  # Lv, x, y
        tc.get("note", "")
    ]
    # LD cases use separate color key
    if tc["Local Dimming"] != "关":
        fill_key = (tc["图像模式"], "LD")
    else:
        fill_key = (tc["图像模式"], tc["hdr_sdr"])
    row_fill = mode_hdr_fills[fill_key]
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.alignment = center
        c.border = thin_border
        c.fill = row_fill

ws.freeze_panes = "A3"


# ===== Sheet 2: Test plan =====
ws2 = wb.create_sheet("测试计划")

plan_items = [
    ("测试目标", "探究图像模式、小窗口大小、峰值亮度、白块亮度、HDR/SDR对Boost峰值亮度的影响规律"),
    ("核心问题1", "不同图像模式(鲜艳/标准/影院)下，各峰值亮度档位的boost提升量是否有差异？"),
    ("核心问题2", "不同白块亮度(100/200/600nit)下boost行为是否不同？低亮度下boost是否失效？"),
    ("核心问题3", "HDR vs SDR下boost行为是否不同？低亮度SDR下boost是否同样失效？"),
    ("信号源", "MURIDEO 8K SEVEN GENERATOR"),
    ("测量设备", "CA-410 色彩分析仪"),
    ("HDR测试", "图像模式: 鲜艳/标准/影院, 窗口: 10%/50%/100%, 峰值亮度: 关/弱/中/强, 白块: 100/200/600nit\n3×3×4×3 = 108组"),
    ("SDR测试", "图像模式: 鲜艳/标准/影院, 窗口: 10%/50%, 峰值亮度: 关/弱/中/强, 白块: 100nit\n3×2×4×1 = 24组"),
    ("LD×Boost测试", "图像模式: 鲜艳/标准/影院, 窗口: 10%/50%, 峰值亮度=强, LD: 弱/中/强, 白块: 100nit, HDR\n3×2×3 = 18组 (LD=关已含在HDR组中)"),
    ("总计", "108 + 24 + 18 = 150组"),
    ("固定条件", "当前背光值=100; HDR/SDR组: LD=关; LD×Boost组: 峰值亮度=强, 白块亮度=100nit, HDR"),
    ("每组测量次数", "1次"),
    ("操作要点1", "HDR时MURIDEO输出HDR10/PQ EOTF/BT.2020格式"),
    ("操作要点2", "SDR时MURIDEO输出BT.709/Gamma 2.2格式"),
    ("操作要点3", "白块亮度通过MURIDEO调节，电视端设置不变"),
    ("操作要点4", "每次切换参数后等待3秒，确保屏幕稳定"),
    ("操作要点5", "保持环境光恒定（暗室或固定照度）"),
    ("操作要点6", "CA-410探头对准MURIDEO输出白块中心位置"),
    ("操作要点7", "关注色度变化：boost可能改变x/y值，不仅看Lv"),
    ("行颜色说明", "HDR: 橙色=鲜艳, 绿色=标准, 蓝色=影院\nSDR: 浅黄=鲜艳, 浅绿=标准, 浅蓝=影院\nLD: 深橙=鲜艳, 深绿=标准, 深蓝=影院"),
    ("预期输出", "各图像模式的boost增益对比、HDR/SDR差异、白块亮度阈值、窗口大小与boost关系曲线"),
]

for i, (k, v) in enumerate(plan_items, 1):
    ck = ws2.cell(row=i, column=1, value=k)
    ck.font = Font(bold=True, size=11)
    cv = ws2.cell(row=i, column=2, value=v)
    cv.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[i].height = 30 if "\n" in v else 20


# ===== Sheet 3: Factor levels =====
ws3 = wb.create_sheet("因子水平")

ref_headers = ["因子", "水平", "类型", "说明", "来源"]
for i, h in enumerate(ref_headers, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

ref_data = [
    ("A: 图像模式", "鲜艳、标准、影院", "离散", "显示预设模式，不同模式对boost增益可能不同", "电视端"),
    ("B: 峰值亮度", "关、弱、中、强", "有序离散", "Boost峰值亮度档位", "电视端"),
    ("C: 当前背光值", "100 (固定)", "连续", "背光PWM/亮度设置值，本次测试固定为100", "电视端"),
    ("D: Local Dimming", "关、弱、中、强", "有序离散", "HDR/SDR组固定为关；LD×Boost组测试弱/中/强与boost强的关系", "电视端"),
    ("E: 小窗口大小", "HDR: 10%/50%/100%, SDR: 10%/50%", "连续", "信号源输出测试图案窗口大小", "MURIDEO"),
    ("F: HDR/SDR", "HDR、SDR", "离散", "HDR: HDR10/PQ/BT.2020; SDR: BT.709/Gamma 2.2", "MURIDEO"),
    ("G: 白块亮度", "HDR: 100/200/600nit, SDR: 100nit", "连续", "信号源输出白块亮度，100nit=极低亮度片源", "MURIDEO"),
]
for r, (factor, levels, ftype, desc, source) in enumerate(ref_data, 2):
    for c_idx, val in enumerate([factor, levels, ftype, desc, source], 1):
        c = ws3.cell(row=r, column=c_idx, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if source == "MURIDEO":
            c.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")


# ===== Sheet 4: MURIDEO setup reference =====
ws4 = wb.create_sheet("MURIDEO设置参考")

setup_headers = ["测试编号段", "小窗口大小", "HDR/SDR", "白块亮度(nit)", "备注"]
for i, h in enumerate(setup_headers, 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.font = header_font_white
    c.fill = murideo_fill
    c.alignment = center
    c.border = thin_border

# Group by image mode + window size + hdr_sdr
setup_data = []

# HDR groups
for mode in IMAGE_MODES:
    for win in WINDOW_SIZES:
        start_idx = None
        end_idx = None
        for i, tc in enumerate(test_cases):
            if tc["图像模式"] == mode and tc["窗口大小"] == f"{win}%" and tc["hdr_sdr"] == "HDR":
                num = int(tc["编号"][1:])
                if start_idx is None:
                    start_idx = num
                end_idx = num
        setup_data.append((f"T{start_idx:03d}-T{end_idx:03d}", f"{win}%", "HDR", "100/200/600", f"{mode}模式"))

setup_data.append(("", "", "", "", ""))

# SDR groups
for mode in IMAGE_MODES:
    for win in SDR_WINDOW_SIZES:
        start_idx = None
        end_idx = None
        for i, tc in enumerate(test_cases):
            if tc["图像模式"] == mode and tc["窗口大小"] == f"{win}%" and tc["hdr_sdr"] == "SDR":
                num = int(tc["编号"][1:])
                if start_idx is None:
                    start_idx = num
                end_idx = num
        setup_data.append((f"T{start_idx:03d}-T{end_idx:03d}", f"{win}%", "SDR", "100", f"{mode}模式"))

setup_data.append(("", "", "", "", ""))

# LD×Boost groups
for mode in IMAGE_MODES:
    start_idx = None
    end_idx = None
    for i, tc in enumerate(test_cases):
        if (tc["图像模式"] == mode and tc["Local Dimming"] != "关"
                and tc["峰值亮度"] == "强" and tc["hdr_sdr"] == "HDR"):
            num = int(tc["编号"][1:])
            if start_idx is None:
                start_idx = num
            end_idx = num
    setup_data.append((f"T{start_idx:03d}-T{end_idx:03d}", "10%/50%", "HDR", "100", f"{mode}模式 LD×Boost强"))

setup_data.append(("", "", "", "", ""))
setup_data.append(("通用设置-HDR", "", "", "", ""))
setup_data.append(("分辨率", "3840×2160", "", "", "4K匹配电视"))
setup_data.append(("色深", "10bit", "", "", "避免色带影响测量"))
setup_data.append(("HDR格式", "HDR10", "", "", "HDR模式"))
setup_data.append(("EOTF", "PQ (ST.2084)", "", "", "HDR10标准EOTF"))
setup_data.append(("色域", "BT.2020", "", "", "匹配电视广色域模式"))
setup_data.append(("", "", "", "", ""))
setup_data.append(("通用设置-SDR", "", "", "", ""))
setup_data.append(("分辨率", "3840×2160", "", "", "4K匹配电视"))
setup_data.append(("色深", "10bit", "", "", "避免色带影响测量"))
setup_data.append(("EOTF", "Gamma 2.2", "", "", "SDR标准"))
setup_data.append(("色域", "BT.709", "", "", "SDR标准色域"))

for r, vals in enumerate(setup_data, 2):
    for c_idx, val in enumerate(vals, 1):
        c = ws4.cell(row=r, column=c_idx, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if val in ("通用设置-HDR", "通用设置-SDR"):
            c.font = Font(bold=True, size=11)

# Auto-fit column widths for all sheets
def autofit_columns(sheet):
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val = str(cell.value)
                # Handle multiline: take the longest line
                line_max = max(len(line) for line in val.split('\n'))
                # CJK characters are wider, estimate ~2x
                cjk_len = sum(1 for ch in val if '一' <= ch <= '鿿' or '　' <= ch <= '〿')
                adjusted = line_max + cjk_len
                if adjusted > max_len:
                    max_len = adjusted
        if max_len > 0:
            sheet.column_dimensions[col_letter].width = max_len + 4  # padding

for sheet in wb.worksheets:
    autofit_columns(sheet)

# Save
output = "/home/liaoxiaolan/claudeCodeProject/read410/Boost峰值亮度测试用例.xlsx"
wb.save(output)
print(f"Saved to: {output}")
print(f"Total test cases: {len(test_cases)}")
